# Define DryBean dataset for training
# Author: wyh
# Date: 2024.1.27

import numpy as np
import openpyxl
import numpy as np
from torch.utils.data import Dataset
from sklearn.preprocessing import LabelEncoder


def load_drybean_dataset(data_dir):
    # data_dir = "/data2/wyh-dataset/tabular-dataset/DryBeanDataset/Dry_Bean_Dataset.xlsx"
    data_dir = data_dir + '/DryBeanDataset/Dry_Bean_Dataset.xlsx'
    readbook = openpyxl.load_workbook(data_dir)
    sheet = readbook['Dry_Beans_Dataset']

    # 数据集中数据的总数量
    n_samples = sheet.max_row - 1
    # 数据集中特征的种类个数
    n_features = sheet.max_column - 1
    # empty()函数构造一个未初始化的矩阵，行数为数据集数量，列数为特征值的种类个数
    data = np.empty((n_samples, n_features))

    # empty()函数构造一个未初始化的矩阵，行数为数据集数量，1列
    labels = np.empty((n_samples,), dtype=np.int64)

    index = 0
    for i in sheet.values:
        if index != 0:
            data[index - 1] = np.asarray(i[0:-1], dtype=np.object_)
            index += 1
        else:
            index += 1

    # 读取最后一列的标签数据
    label_data = [i[-1] for i in sheet.values][1:]
    # 使用LabelEncoder对标签进行编码
    label_encoder = LabelEncoder()
    labels = label_encoder.fit_transform(label_data)

    # print(data.shape)# (13611, 16)
    # print(labels.shape)# (13611,)
    return data, labels


class DryBeanDataset(Dataset):
    def __init__(self, data_dir):
        self.data, self.labels = load_drybean_dataset(data_dir)

    def __getitem__(self, index):
        return self.data[index], self.labels[index]

    def __len__(self):
        return len(self.data)
    
class DryBeanTrainDataset(Dataset):
    def __init__(self, noise_ratio, data, labels, train_indexes=None, transform=None, target_transform=None):
        # define the number of classes
        self.num_classes = 7

        # define data
        self.data = data
        self.labels = labels

        # define the ratio
        self.noise_ratio = noise_ratio
        self.transform=transform
        self.target_transform=target_transform

        # choose the data and label in dataset
        if train_indexes is not None:
            self.train_data = np.array(self.data)[train_indexes]
            self.true_labels = np.array(self.labels)[train_indexes]
            self.train_labels = np.array(self.labels)[train_indexes]
        else:
            self.train_data = np.array(self.data)
            self.true_labels = np.array(self.labels)
            self.train_labels = np.array(self.labels)
        self.soft_labels = np.zeros((len(self.train_labels), self.num_classes), dtype=np.float32)
        self.noise_or_not = [False] * len(self.train_data)  # 初始化为全部为False

    def __len__(self):
        return len(self.train_labels)
    
    # make symmetric noise labels
    # The difference between generating random noise overall and generating noise for each class is not significant
    def symmetric_noise(self):
        # generate random noise across the dataset
        idxes = np.random.permutation(len(self.true_labels))
        noise_num = int(len(self.true_labels) * self.noise_ratio)
        for i in range(len(idxes)):
            if i < noise_num:
                # a randomly generated label that differs from the privious label
                exclude_class = self.true_labels[idxes[i]]
                label_sym = np.random.choice(np.delete(np.arange(self.num_classes), exclude_class), 1)[0]
                self.train_labels[idxes[i]] = label_sym
                self.noise_or_not[idxes[i]] = True
            self.soft_labels[idxes[i]][self.train_labels[idxes[i]]] = 1
    
    # make asymmetric noise labels
    def asymmetric_noise(self):
        dic = {0:1, 1:2, 2:3, 3:4, 4:5, 5:6, 6:0}
        for i in range(self.num_classes):
            idxes = np.where(self.true_labels == i)[0]
            np.random.shuffle(idxes)
            noise_num = int(int(len(idxes)) * self.noise_ratio) #noise_ratio
            for j in range(len(idxes)):
                if j< noise_num and i in dic:
                    self.train_labels[idxes[j]] = dic[i]
                    self.noise_or_not[idxes[j]] = True
                self.soft_labels[idxes[j]][self.train_labels[idxes[j]]] = 1
 
    def __getitem__(self, index):
        data, label, true_label, soft_label = self.train_data[index], self.train_labels[index], self.true_labels[index], self.soft_labels[index]
        
        # 判断是否是噪音数据
        is_noise = False
        if label != true_label:
            is_noise = True
            self.noise_or_not[index] = True  # 更新对应位置的noise_or_not元素
        
        
        if self.transform is not None:
            data = self.transform(data)

        if self.target_transform is not None:
            label = self.target_transform(label)

        return data, label, true_label, is_noise, soft_label, index

class DryBeanValDataset(Dataset):
    def __init__(self, data, labels, val_indexes, transform=None, target_transform=None):
        self.transform = transform
        self.target_transform = target_transform
        self.data = data
        self.labels = labels
        self.val_data = np.array(self.data)[val_indexes]
        self.val_labels = np.array(self.labels)[val_indexes]
    
    def __len__(self):
        return len(self.val_labels)
    
    def __getitem__(self, index):  
        data, label = self.val_data[index], self.val_labels[index]
        # doing this so that it is consistent with all other datasets.

        if self.transform is not None:
            data = self.transform(data)

        if self.target_transform is not None:
            label = self.target_transform(label)

        return data, label

def classwise_split(dataset_targets, ratio, num_classes):
    # select {ratio*len(labels)} images from the images
    train_val_label = np.array(dataset_targets)
    train_indexes = []
    val_indexes = []

    for id in range(num_classes):
        indexes = np.where(train_val_label == id)[0]
        # print(len(indexes))
        np.random.shuffle(indexes)
        train_num = int(len(indexes)*ratio)
        train_indexes.extend(indexes[:train_num])
        val_indexes.extend(indexes[train_num:])
    
    np.random.shuffle(train_indexes)
    np.random.shuffle(val_indexes)

    return train_indexes, val_indexes

# data_dir = "/data2/wyh-dataset/tabular-dataset"
# dataset = DryBeanDataset(data_dir=data_dir)

# train_val_indexes, test_indexes = classwise_split(dataset_targets=dataset.labels, ratio=0.8, num_classes=7)
# train_val_data = np.array(dataset.data)[train_val_indexes]
# train_val_labels = np.array(dataset.labels)[train_val_indexes]
# train_indexes, val_indexes = classwise_split(dataset_targets=train_val_labels, ratio=0.9, num_classes=7)

# train_dataset = DryBeanTrainDataset(data=train_val_data, labels=train_val_labels ,noise_ratio=0.6, train_indexes=train_indexes)
# val_dataset = DryBeanValDataset(data = train_val_data, labels=train_val_labels, val_indexes=val_indexes)
# test_dataset = DryBeanValDataset(data = dataset.data, labels= dataset.labels, val_indexes=test_indexes)
# print(len(train_dataset))
# train_dataset.asymmetric_noise()
# print(train_dataset.noise_or_not.count(True))
# print(len(val_dataset))

